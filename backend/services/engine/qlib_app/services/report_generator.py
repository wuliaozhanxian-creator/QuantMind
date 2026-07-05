"""回测报告生成服务"""

import io
import logging
from typing import Any, Optional

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

matplotlib.use("Agg")  # 非 GUI 后端

from backend.services.engine.qlib_app.utils.structured_logger import (
    StructuredTaskLogger,
)

logger = logging.getLogger(__name__)
task_logger = StructuredTaskLogger(logger, "ReportGenerator")

class PDFReportGenerator:
    """PDF 回测报告生成器"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            "CustomTitle",
            parent=self.styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1a1a1a"),
            spaceAfter=30,
            alignment=1,  # 居中
        )
        self.heading_style = ParagraphStyle(
            "CustomHeading",
            parent=self.styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#333333"),
            spaceAfter=12,
        )

    def generate(self, backtest_result: dict[str, Any]) -> bytes:
        """
        生成 PDF 报告

        Args:
            backtest_result: 回测结果数据

        Returns:
            PDF 文件字节流
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        # 1. 标题
        story.append(Paragraph("量化回测报告", self.title_style))
        story.append(Spacer(1, 0.2 * inch))

        # 2. 回测概览
        story.extend(self._build_overview_section(backtest_result))
        story.append(Spacer(1, 0.3 * inch))

        # 3. 核心指标表格
        story.extend(self._build_metrics_table(backtest_result))
        story.append(Spacer(1, 0.3 * inch))

        # 4. 权益曲线图
        equity_img = self._generate_equity_curve(backtest_result)
        if equity_img:
            story.append(Paragraph("权益曲线", self.heading_style))
            story.append(equity_img)
            story.append(Spacer(1, 0.2 * inch))

        # 5. 回撤曲线图
        drawdown_img = self._generate_drawdown_curve(backtest_result)
        if drawdown_img:
            story.append(Paragraph("回撤曲线", self.heading_style))
            story.append(drawdown_img)
            story.append(Spacer(1, 0.2 * inch))

        # 6. 月度收益热力图
        monthly_img = self._generate_monthly_returns(backtest_result)
        if monthly_img:
            story.append(PageBreak())
            story.append(Paragraph("月度收益率", self.heading_style))
            story.append(monthly_img)
            story.append(Spacer(1, 0.2 * inch))

        # 7. 持仓分布
        position_img = self._generate_position_distribution(backtest_result)
        if position_img:
            story.append(Paragraph("持仓分布（最后一日）", self.heading_style))
            story.append(position_img)

        # 构建 PDF
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        return pdf_bytes

    def _build_overview_section(self, result: dict[str, Any]) -> list:
        """构建概览部分"""
        config = result.get("config", {})

        overview_data = [
            ["回测 ID", result.get("backtest_id", "N/A")],
            ["策略类型", config.get("strategy_type", "N/A")],
            [
                "回测时间",
                f"{config.get('start_date', 'N/A')} 至 {config.get('end_date', 'N/A')}",
            ],
            ["初始资金", f"{config.get('initial_capital', 0):,.0f} 元"],
            ["基准指数", config.get("benchmark", "N/A")],
            ["股票池", config.get("universe", "N/A")],
            [
                "创建时间",
                (
                    result.get("created_at", "N/A")[:19]
                    if result.get("created_at")
                    else "N/A"
                ),
            ],
        ]

        table = Table(overview_data, colWidths=[2 * inch, 4 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f0f0")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )

        return [Paragraph("回测概览", self.heading_style), table]

    def _build_metrics_table(self, result: dict[str, Any]) -> list:
        """构建指标表格"""
        metrics_data = [
            ["指标", "数值"],
            ["年化收益率", f"{result.get('annual_return', 0) * 100:.2f}%"],
            ["总收益率", f"{result.get('total_return', 0) * 100:.2f}%"],
            ["夏普比率", f"{result.get('sharpe_ratio', 0):.3f}"],
            ["最大回撤", f"{result.get('max_drawdown', 0) * 100:.2f}%"],
            ["波动率", f"{result.get('volatility', 0) * 100:.2f}%"],
            ["Alpha", f"{result.get('alpha', 0):.3f}"],
            ["Beta", f"{result.get('beta', 0):.3f}"],
            ["信息比率", f"{result.get('information_ratio', 0):.3f}"],
            ["基准收益率", f"{result.get('benchmark_return', 0) * 100:.2f}%"],
        ]

        table = Table(metrics_data, colWidths=[3 * inch, 3 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )

        return [Paragraph("核心指标", self.heading_style), table]

    def _generate_equity_curve(self, result: dict[str, Any]) -> Image | None:
        """生成权益曲线图"""
        equity_curve = result.get("equity_curve")
        if not equity_curve:
            return None

        try:
            df = pd.DataFrame(equity_curve)
            if "date" not in df.columns or "value" not in df.columns:
                return None

            fig, ax = plt.subplots(figsize=(8, 4))
            ax.plot(
                pd.to_datetime(df["date"]), df["value"], linewidth=2, color="#4472C4"
            )
            ax.set_xlabel("日期", fontsize=10)
            ax.set_ylabel("权益", fontsize=10)
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()

            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format="png", dpi=150)
            plt.close(fig)
            img_buffer.seek(0)

            return Image(img_buffer, width=6 * inch, height=3 * inch)
        except Exception as e:
            task_logger.error(
                "generate_equity_curve_failed", "生成权益曲线图失败", error=str(e)
            )
            return None

    def _generate_drawdown_curve(self, result: dict[str, Any]) -> Image | None:
        """生成回撤曲线图"""
        drawdown_curve = result.get("drawdown_curve")
        if not drawdown_curve:
            return None

        try:
            df = pd.DataFrame(drawdown_curve)
            if "date" not in df.columns or "drawdown" not in df.columns:
                return None

            fig, ax = plt.subplots(figsize=(8, 4))
            ax.fill_between(
                pd.to_datetime(df["date"]),
                df["drawdown"] * 100,
                0,
                color="#C55A11",
                alpha=0.3,
            )
            ax.plot(
                pd.to_datetime(df["date"]),
                df["drawdown"] * 100,
                linewidth=2,
                color="#C55A11",
            )
            ax.set_xlabel("日期", fontsize=10)
            ax.set_ylabel("回撤 (%)", fontsize=10)
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()

            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format="png", dpi=150)
            plt.close(fig)
            img_buffer.seek(0)

            return Image(img_buffer, width=6 * inch, height=3 * inch)
        except Exception as e:
            task_logger.error(
                "generate_drawdown_curve_failed", "生成回撤曲线图失败", error=str(e)
            )
            return None

    def _generate_monthly_returns(self, result: dict[str, Any]) -> Image | None:
        """生成月度收益热力图"""
        equity_curve = result.get("equity_curve")
        if not equity_curve:
            return None

        try:
            df = pd.DataFrame(equity_curve)
            if "date" not in df.columns or "value" not in df.columns:
                return None

            df["date"] = pd.to_datetime(df["date"])
            df["return"] = df["value"].pct_change()
            df["year"] = df["date"].dt.year
            df["month"] = df["date"].dt.month

            # 计算月度收益
            monthly = (
                df.groupby(["year", "month"])["return"].sum().unstack(fill_value=0)
            )

            fig, ax = plt.subplots(figsize=(10, 4))
            im = ax.imshow(monthly.values * 100, cmap="RdYlGn", aspect="auto")
            ax.set_xticks(np.arange(12))
            ax.set_xticklabels([f"{i + 1}月" for i in range(12)])
            ax.set_yticks(np.arange(len(monthly)))
            ax.set_yticklabels(monthly.index)
            plt.colorbar(im, ax=ax, label="收益率 (%)")
            plt.tight_layout()

            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format="png", dpi=150)
            plt.close(fig)
            img_buffer.seek(0)

            return Image(img_buffer, width=6 * inch, height=3 * inch)
        except Exception as e:
            task_logger.error(
                "generate_monthly_returns_failed",
                "生成月度收益热力图失败",
                error=str(e),
            )
            return None

    def _generate_position_distribution(self, result: dict[str, Any]) -> Image | None:
        """生成持仓分布饼图"""
        positions = result.get("positions")
        if not positions or len(positions) == 0:
            return None

        try:
            # 取前 10 大持仓
            top_positions = sorted(
                positions, key=lambda x: x.get("weight", 0), reverse=True
            )[:10]
            labels = [p.get("instrument", "Unknown") for p in top_positions]
            sizes = [p.get("weight", 0) * 100 for p in top_positions]

            fig, ax = plt.subplots(figsize=(8, 6))
            ax.pie(sizes, labels=labels, autopct="%1.1f%%", startangle=90)
            ax.axis("equal")
            plt.tight_layout()

            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format="png", dpi=150)
            plt.close(fig)
            img_buffer.seek(0)

            return Image(img_buffer, width=5 * inch, height=4 * inch)
        except Exception as e:
            task_logger.error(
                "generate_position_distribution_failed",
                "生成持仓分布图失败",
                error=str(e),
            )
            return None

class ExcelReportGenerator:
    """Excel 回测报告生成器"""

    def generate(self, backtest_result: dict[str, Any]) -> bytes:
        """
        生成 Excel 报告

        Args:
            backtest_result: 回测结果数据

        Returns:
            Excel 文件字节流
        """
        wb = Workbook()
        wb.remove(wb.active)  # 移除默认 sheet

        # Sheet 1: 核心指标
        self._create_metrics_sheet(wb, backtest_result)

        # Sheet 2: 权益曲线
        self._create_equity_sheet(wb, backtest_result)

        # Sheet 3: 交易明细
        self._create_trades_sheet(wb, backtest_result)

        # Sheet 4: 持仓明细
        self._create_positions_sheet(wb, backtest_result)

        # Sheet 5: 每日收益率
        self._create_daily_returns_sheet(wb, backtest_result)

        # 保存到字节流
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        return excel_bytes

    def _create_metrics_sheet(self, wb: Workbook, result: dict[str, Any]):
        """创建核心指标 Sheet"""
        ws = wb.create_sheet("核心指标")

        # 标题样式
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        # 写入数据
        ws.append(["指标", "数值"])
        ws.append(["年化收益率", f"{result.get('annual_return', 0) * 100:.2f}%"])
        ws.append(["总收益率", f"{result.get('total_return', 0) * 100:.2f}%"])
        ws.append(["夏普比率", f"{result.get('sharpe_ratio', 0):.3f}"])
        ws.append(["最大回撤", f"{result.get('max_drawdown', 0) * 100:.2f}%"])
        ws.append(["波动率", f"{result.get('volatility', 0) * 100:.2f}%"])
        ws.append(["Alpha", f"{result.get('alpha', 0):.3f}"])
        ws.append(["Beta", f"{result.get('beta', 0):.3f}"])
        ws.append(["信息比率", f"{result.get('information_ratio', 0):.3f}"])
        ws.append(["基准收益率", f"{result.get('benchmark_return', 0) * 100:.2f}%"])

        # 应用样式
        ws["A1"].fill = header_fill
        ws["B1"].fill = header_fill
        ws["A1"].font = header_font
        ws["B1"].font = header_font

        # 调整列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15

    def _create_equity_sheet(self, wb: Workbook, result: dict[str, Any]):
        """创建权益曲线 Sheet"""
        ws = wb.create_sheet("权益曲线")
        equity_curve = result.get("equity_curve", [])

        if equity_curve:
            df = pd.DataFrame(equity_curve)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # 标题样式
            for cell in ws[1]:
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

    def _create_trades_sheet(self, wb: Workbook, result: dict[str, Any]):
        """创建交易明细 Sheet"""
        ws = wb.create_sheet("交易明细")
        trades = result.get("trades", [])

        if trades:
            df = pd.DataFrame(trades)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # 标题样式
            for cell in ws[1]:
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

    def _create_positions_sheet(self, wb: Workbook, result: dict[str, Any]):
        """创建持仓明细 Sheet"""
        ws = wb.create_sheet("持仓明细")
        positions = result.get("positions", [])

        if positions:
            df = pd.DataFrame(positions)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # 标题样式
            for cell in ws[1]:
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

    def _create_daily_returns_sheet(self, wb: Workbook, result: dict[str, Any]):
        """创建每日收益率 Sheet"""
        ws = wb.create_sheet("每日收益率")
        equity_curve = result.get("equity_curve", [])

        if equity_curve:
            df = pd.DataFrame(equity_curve)
            if "date" in df.columns and "value" in df.columns:
                df["daily_return"] = df["value"].pct_change()
                df = df[["date", "daily_return"]].dropna()

                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                # 标题样式
                for cell in ws[1]:
                    cell.fill = PatternFill(
                        start_color="4472C4", end_color="4472C4", fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="FFFFFF")
